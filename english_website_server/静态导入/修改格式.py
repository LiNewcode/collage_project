import re
import time

from docx import Document
import docx
import openpyxl
from openpyxl import load_workbook

from end_of_term_work.Tool.link_db import db_obj


def readExcl(path):
    paragraph = []
    # excel
    wb = load_workbook(path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet:
            row_content = ''
            for cell in row:
                if cell.value is not None:
                    row_content += ' ' + str(cell.value)
            paragraph.append(row_content + '\n')

            # print(row_content)
    content = ''.join((''.join(paragraph)).split('淘宝店铺：https://shop249445206.taobao.com/  掌柜旺旺：新一文化'))
    return content


# 去除页码
def delIndex(content):
    content = ''.join(re.split('\s\d{1,2}\n{1,4}', content))
    print(re.split('\n\s{1,}(\n{1,3})', content))
    content = ''.join(re.split('\n\s{1,}(\n{1,3})', content))
    return content


#   修改标题
def changeTitle(content):
    # need to change:
    s1 = '一、文章题材结构分析'
    s2 = '二、试题解析'
    for i in re.findall('一、.*?\n', content):
        print(i)
    content = re.sub('一、文章.*?\n', '一、文章题材结构分析\n', content)
    content = re.sub('\n.*?、.*?试题.*?\n', '\n二、试题解析\n', content)
    for i in re.findall('\n.*?试题.*?\n', content):
        print(i)
    return content


# 修改题号
def changeQuestionId(content):
    for i in re.findall('\d{1,2} ．|\d{1,2}．|\d{1,2} \.|\d{1,2} 、|\d{1,2}、|\d{1,2}）|\d{1,2} ）',content):
        rep = re.findall('\d{1,2}',i)[0]+'.'
        print(i)
        print(f'repl,{rep}')
        # if ')' in i:
        #     orginal = r'\\)'.join(i.split(')'))
        #     print(orginal)
        #     content = re.sub(orginal, rep, content)
        content = re.sub(i,rep,content)

    return content

#查找答案选项
def findChoice(content):
    for i in re.findall('\n\s{0,}\d{1,2}\..*?【答案】',content,flags=re.S):
        print(i)
# ' ．'
# '．'
if __name__ == '__main__':
    # year = '2015'
    # content = readExcl('F:\\PyCourseDemo\\end_of_term_work\\静态导入\\excel\\2010-2020解析.xlsx')
    # print(content)
    with open('F:\\PyCourseDemo\\end_of_term_work\\静态导入\\excel\\2010-2020_2.txt', 'r', encoding='utf-8') as fp:
        content = fp.read()
        # findChoice(content)
        content = re.split('(\d{4}年全国硕士研究生招生考试英语（一）答案详解)',content)
        print(content)
        index = 2010
        for i in range(2,24,2):
        # print(content)
            print(content[i])
            with open(f'F:\\PyCourseDemo\\end_of_term_work\\静态导入\\excel\\{str(index)}.txt','w',encoding='utf-8')as fx:
                fx.write(content[i])
            index+=1
    #  changeTitle(content=content)
    # print(content)
    # with open('F:\\PyCourseDemo\\end_of_term_work\\静态导入\\excel\\test.txt', 'r', encoding='utf-8') as fp:
    #     content = fp.read()
    #     print(re.findall('\d{1,2}\.(.*?)【答案】',content,flags=re.S))
    # with open('F:\\PyCourseDemo\\end_of_term_work\\静态导入\\excel\\2010-2020_2.txt', 'w', encoding='utf-8') as fp:
    #     # fp.write(delIndex(content))
    #     # fp.write(changeTitle(content))
    #     # fp.write(changeQuestionId(content=content))