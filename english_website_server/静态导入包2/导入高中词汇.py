"""
    将高中3500词插入
    date 2022.02.26
    author lyz
"""
import re
import time

import openpyxl
from openpyxl import load_workbook

from end_of_term_work.Tool.link_db import db_obj

if __name__ == '__main__':
    path = '六级词汇.xlsx'
    wb = load_workbook(path)
    final_wb = openpyxl.Workbook()

    start_sheet = wb.active
    wor_dic ={}
    db = db_obj(col='lj')
    start = time.time()
    # print(db.dropCol())
    wor_list = []
    for i in range(2, start_sheet.max_row + 1):
        word = start_sheet['A' + str(i)].value
        # if len(re.findall(r'\#|\(|\)|\;|\:|\"|\&|\^|\!|\=|‘|“|-{2,}|/|\.',str(word)))>0:
        #     continue
        wor_list.append(str(word).strip())
        # phs = start_sheet['B' + str(i)].value
        # explain = start_sheet['C' + str(i)].value
    obj = {'word': wor_list}
    try:
        db.insert(obj)
    except:
        print(obj)
    # for i in db.search():
    #     print(i)
    print(len(wor_list))
    end = time.time()
    print(f"spend time :[ {end - start} s]")