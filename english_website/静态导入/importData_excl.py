"""
    将考研5500词插入
    date 2021.11.27
    author lyz
"""
import time

import openpyxl
from openpyxl import load_workbook

from end_of_term_work.Tool.link_db import db_obj

if __name__ == '__main__':
    path = 'F:\英语\考研英语词.xlsx'
    wb = load_workbook(path)
    final_wb = openpyxl.Workbook()

    start_sheet = wb.active
    wor_dic ={}
    db = db_obj(col='e_pee')
    start = time.time()
    print(db.dropCol())
    for i in range(2, start_sheet.max_row + 1):
        word = start_sheet['A' + str(i)].value
        phs = start_sheet['B' + str(i)].value
        explain = start_sheet['C' + str(i)].value
        obj = {'word':[word,phs,explain]}
        try:
            db.insert(obj)
        except:
            print(obj)
    for i in db.search():
        print(i)

    end = time.time()
    print(f"spend time :[ {end - start} s]")