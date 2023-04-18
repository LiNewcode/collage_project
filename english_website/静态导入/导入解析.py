import json
import re
import time

from docx import Document
import docx
import openpyxl
from openpyxl import load_workbook

from end_of_term_work.Tool.link_db import db_obj

index_question = 1
ans_list = []
pee_type = None
year = None
wenZhang = None

def deal_trans(ftr):
    ftr = ftr.split('\n')
    new_ftr = []
    for i in ftr:
        if (i.strip() != '' and i.split() is not None):
            new_ftr.append(i)
    index = 0
    for i in new_ftr:
        sentence = re.split('\s', i)
        sentence = ''.join(sentence)
        new_ftr[index] = sentence
        index += 1
    return new_ftr
    # print(new_ftr)


def res_cell(cell, pc=None):
    if pc:
        point = (re.findall('【考点】(.*?)【解析】', cell, re.S)[0]).strip()
        analysis = (re.findall('【解析】(.*?)【词汇】', cell, re.S)[0]).strip()
        word = (re.findall('【词汇】(.*?)【参考译文】', cell, re.S)[0]).strip()
        trans = (re.findall('【参考译文】(.*)', cell, re.S)[0]).strip()
        print(trans)
        trans = ''.join(trans.split('\n'))
        return point, analysis, word, trans
    ans = ''
    ans = (re.findall('【答案】(.*?)\n', cell)[0]).strip()
    point = ''
    point = (re.findall('【考点】(.*?)\n', cell)[0]).strip()
    RightRate = 0
    analysis = ''
    analysis = (re.split('【解析】', cell)[1]).strip()
    analysis = ' '.join(re.split('\n', analysis))
    return ans, point, RightRate, analysis


def res_txt(content, count, pc=None):
    global index_question
    global ans_list
    global year
    global pee_type
    sms = content.split('一、文章题材结构分析')[1].split('二、试题解析')[0]
    sms = sms.strip().split('\n')
    qr = content.split('二、试题解析')[1].split('三、全文翻译')[0]
    ftr = content.split('三、全文翻译')[1]
    ftr = deal_trans(ftr)
    index = 1
    # db = db_obj(col='peeQrSection')
    # res = ''
    # for i in re.split(r'-[0-9]{1,2}\.', qr)[1:]:
    #     if index == count:
    #         break
    #     if pc:
    #         _id = f'{pee_type}-{str(year)}-{str(index_question)}'
    #         point, analysis, word, trans = res_cell(i, pc)
    #         db.update({'_id': _id}, {'ans': trans, 'point': point, 'quanZhanZuoDa': 0, 'rightRate': 0,
    #                                  'analysis': [word, analysis]})
    #     else:
    #         ans, point, RightRate, analysis = res_cell(i)
    #         ans = ans_list[index_question - 1]
    #         _id = f'{pee_type}-{str(year)}-{str(index_question)}'
    #         db.update({'_id': _id}, {'ans': ans, 'point': point, 'quanZhanZuoDa': 0, 'rightRate': RightRate,
    #                                  'analysis': analysis})
    #         # res += f'[year:{year}][index:{index}]\n[ans]:{ans}\n[point]:{point}\n[analysis]:{analysis}\n'
    #
    #     index_question += 1
    #
    #     index += 1
    # return res
    # db.close()
    global wenZhang
    _id = f'{pee_type}-{str(year)}-{wenZhang}'
    db = db_obj(col='peeQrWhole')
    db.update({'_id': _id}, {'sms': sms, 'ftr': ftr})
    db.close()


def sec_1(content, year):
    global index_question
    global wenZhang
    wenZhang = 'sec_1-a-0'
    index_question = 1
    # question = get_selection(content, 1)
    res_txt(content, 21)


def sec_2_pa_text(content, text_id, question_req, year):
    print(year, text_id)
    res_txt(content, 6)


def sec_2_pa(pa, content, year):
    text = re.split('(Text 1|Text 2|Text 3|Text 4|Text 5)', pa)[1:]
    index = 0
    # print(text)
    # print(len(text))
    global wenZhang
    t=1
    for i in range(1, 8, 2):
        # print(i)
        wenZhang = f'sec_2-a-{t}'
        t+=1
        sec_2_pa_text(text[i].strip(), str(i), None, year)


def sec_2_pb(pb, year):
    global wenZhang
    content = pb
    print(year, ' pb ')
    wenZhang = f'sec_2-b-5'
    res_txt(content, 6)


def sec_2_pc(pc, year):
    global wenZhang
    content = pc
    print(year, ' pc ')
    wenZhang = f'sec_2-c-6'
    res_txt(content, 6, pc=True)


def sec_2(content, year):
    pa = content.split('Part B')[0]
    sec_2_pa(pa, content, year)
    pb = content.split('Part B')[1].split('Part C')[0]
    sec_2_pb(pb, year)
    pc = content.split('Part C')[1]
    sec_2_pc(pc, year)


def sec_3(content, year):
    # examine topic
    et = '一、审题谋篇'
    # Model text
    mt = '二、参考范文'
    # Writing skills
    ws = '三、模板例句'  # '三、写作技巧' or
    pa = content.split('Part B')[0].split('Part A')[1]
    pb = content.split('Part B')[1]
    pa_et = pa.split(mt)[0].split(et)[1]
    pb_et = pb.split(mt)[0].split(et)[1]
    pa_mt = pa.split(mt)[1].split(ws)[0]
    pa_mt = pa_mt.split('\n')
    pb_mt = pb.split(mt)[1].split(ws)[0]
    pb_mt = pb_mt.split('\n')
    pa_ws = pa.split(ws)[1]
    pa_ws = pa_ws.split('\n')
    pb_ws = pb.split(ws)[1]
    pb_ws = pb_ws.split('\n')
    # print([pa_et, pa_mt, pa_ws])
    # print([pb_et, pb_mt, pb_ws])
    # id = year + '-3-a'
    # db = db_obj(col='peeQrWhole')
    # db.update({'_id': id}, {'_id': id, 'et': pa_et, 'mt': pa_mt, 'ws': pa_ws})
    # db.close()
    # id = year + '-3-b'
    # db = db_obj(col='peeQrWhole')
    # db.update({'_id': id}, {'_id': id, 'et': pb_et, 'mt': pb_mt, 'ws': pb_ws})
    # db.close()


def readExcl(path):
    paragraph = []
    # excel
    wb = load_workbook(path)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet:
            row_content = ''
            for cell in row:
                if cell.value is not None:
                    row_content += ' ' + str(cell.value)
            paragraph.append(row_content + '\n')

            # print(row_content)
    content = ''.join((''.join(paragraph)).split('淘宝店铺：https://shop249445206.taobao.com/  掌柜旺旺：新一文化'))
    return content


def getAns(ans):
    global ans_list
    ans_list = []
    questions = ans['Questions']
    # print(questions)
    k = 0
    for item in questions:
        if k == 6:
            break
        # print(item)
        k = k + 1
        Scblt = item['SbjSubContentList']
        for j in Scblt:
            ans_list.append(j['Answer'])
    return ans_list


def enter_e_year(y):
    global pee_type
    global year
    pee_type = '1'
    year = y


if __name__ == '__main__':
    for i in range(2011, 2021):
        year = str(i)
        with open(f'{year}.txt', 'r', encoding='utf-8') as fp:
            content = fp.read()
        with open(f'英一{year}.json', 'r', encoding='utf-8') as fp1:
            ans = json.loads(fp1.read())
        getAns(ans)
        enter_e_year(year)
        # print('ans:', ans)
        section = ['Section I', 'Section II', 'Section III']
        section_1 = content.split(section[1])[0]
        sec_1(section_1, year)

        section_2 = (content.split(section[1])[1]).split(section[2])[0]

        sec_2(section_2, year)
        section_3 = content.split(section[2])[1]
        sec_3(section_3, year)
    # word
    # 打开文件
    # text = ''
    # doc = Document("F:\\英语\\考研题目\\10-20解析\\2008年考研英语真题答案及解析.docx")
    # for paragraph in doc.paragraphs:
    #     print(paragraph.text)
    # print(text)
